from datetime import date, datetime
from io import BytesIO
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.schemas.dashboard_schema import OwnerAnalyticsReport, OwnerDashboardResponse


class ExcelExportService:
    _HEADER_FILL = PatternFill("solid", fgColor="1F6F4A")
    _HEADER_FONT = Font(color="FFFFFF", bold=True)
    _TITLE_FONT = Font(color="0F172A", bold=True, size=16)
    _SUBTITLE_FONT = Font(color="334155", bold=True)
    _BORDER = Border(bottom=Side(style="thin", color="D9E2EC"))
    _CURRENCY_FORMAT = '"$"#,##0.00'
    _PERCENT_FORMAT = '0.00%'

    def build_owner_dashboard_workbook(
        self,
        dashboard: OwnerDashboardResponse,
        from_date: date,
        to_date: date
    ) -> BytesIO:
        workbook = Workbook()

        summary_sheet = workbook.active
        summary_sheet.title = "Resumen"
        summary_sheet.append(["Reporte Tap2Eat"])
        summary_sheet.append(["Restaurante", dashboard.restaurant_id])
        summary_sheet.append(["Fecha inicio", from_date.isoformat()])
        summary_sheet.append(["Fecha fin", to_date.isoformat()])
        summary_sheet.append([])
        summary_sheet.append(["Pedidos", "Valor"])
        summary_sheet.append(["Total de pedidos", dashboard.orders.total_orders])
        summary_sheet.append(["Pedidos cobrables", dashboard.orders.sales_order_count])
        summary_sheet.append(["Ventas totales", dashboard.orders.total_sales])
        summary_sheet.append(["Ventas entregadas", dashboard.orders.delivered_sales])
        summary_sheet.append(["Ticket promedio", dashboard.orders.average_ticket])
        summary_sheet.append(["Pedidos creados", dashboard.orders.created_orders])
        summary_sheet.append(["Pedidos aceptados", dashboard.orders.accepted_orders])
        summary_sheet.append(["Pedidos en preparaci\u00f3n", dashboard.orders.preparing_orders])
        summary_sheet.append(["Pedidos listos", dashboard.orders.ready_orders])
        summary_sheet.append(["Pedidos entregados", dashboard.orders.delivered_orders])
        summary_sheet.append(["Pedidos cancelados", dashboard.orders.cancelled_orders])
        summary_sheet.append([])
        summary_sheet.append(["Cat\u00e1logo", "Valor"])
        summary_sheet.append(["Total de productos", dashboard.catalog.total_products])
        summary_sheet.append(["Productos disponibles", dashboard.catalog.available_products])
        summary_sheet.append(["Productos pausados", dashboard.catalog.paused_products])
        summary_sheet.append(["Productos simples", dashboard.catalog.simple_products])
        summary_sheet.append(["Productos personalizables", dashboard.catalog.customizable_products])
        summary_sheet.append(["Productos destacados", dashboard.catalog.featured_products])
        summary_sheet.append([
            "Productos con horario personalizado",
            dashboard.catalog.products_with_custom_schedule
        ])
        summary_sheet.append(["Total de categor\u00edas", dashboard.catalog.total_categories])
        summary_sheet.append(["Categor\u00edas activas", dashboard.catalog.active_categories])

        status_sheet = workbook.create_sheet("Pedidos por estado")
        self._append_table(
            status_sheet,
            ["Estado", "Total"],
            ([metric.status, metric.total] for metric in dashboard.orders.orders_by_status)
        )

        product_type_sheet = workbook.create_sheet("Productos por tipo")
        self._append_table(
            product_type_sheet,
            ["Tipo", "Total"],
            ([metric.product_type, metric.total] for metric in dashboard.catalog.products_by_type)
        )

        product_status_sheet = workbook.create_sheet("Productos por estado")
        self._append_table(
            product_status_sheet,
            ["Estado", "Total"],
            ([metric.status, metric.total] for metric in dashboard.catalog.products_by_status)
        )

        self._format_workbook(workbook)

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output

    def build_owner_analytics_workbook(
        self,
        report: OwnerAnalyticsReport,
        from_date: date,
        to_date: date
    ) -> BytesIO:
        workbook = Workbook()

        summary_sheet = workbook.active
        summary_sheet.title = "Resumen ejecutivo"
        self._build_summary_sheet(summary_sheet, report, from_date, to_date)

        orders_sheet = workbook.create_sheet("Detalle de pedidos")
        self._append_table(
            orders_sheet,
            [
                "Folio pedido",
                "Fecha",
                "Hora",
                "Sucursal",
                "Cliente",
                "Estado del pedido",
                "M\u00e9todo de pago",
                "Estado de pago",
                "Subtotal",
                "Total",
                "Cantidad recibida",
                "Cambio entregado",
                "Proveedor de pago",
                "Referencia de pago",
                "N\u00famero de productos/items",
                "Notas del pedido",
            ],
            (
                [
                    detail.order_folio,
                    detail.date,
                    detail.time,
                    detail.branch,
                    detail.customer,
                    detail.order_status,
                    detail.payment_method,
                    detail.payment_status,
                    detail.subtotal,
                    detail.total,
                    detail.amount_received,
                    detail.change_amount,
                    detail.payment_provider,
                    detail.payment_reference,
                    detail.items_count,
                    detail.notes,
                ]
                for detail in report.order_details
            )
        )

        sold_products_sheet = workbook.create_sheet("Detalle de productos vendidos")
        self._append_table(
            sold_products_sheet,
            [
                "Folio pedido",
                "Fecha",
                "Sucursal",
                "Producto",
                "Cantidad",
                "Precio unitario",
                "Total del producto",
                "Modificadores o extras",
                "Estado del pedido",
            ],
            (
                [
                    detail.order_folio,
                    detail.date,
                    detail.branch,
                    detail.product,
                    detail.quantity,
                    detail.unit_price,
                    detail.product_total,
                    detail.modifiers,
                    detail.order_status,
                ]
                for detail in report.sold_product_details
            )
        )

        sales_sheet = workbook.create_sheet("Ventas por d\u00eda")
        self._append_table(
            sales_sheet,
            ["Fecha", "Total vendido", "\u00d3rdenes", "Ticket promedio del d\u00eda"],
            (
                [metric.date, metric.total_sales, metric.total_orders, metric.average_ticket]
                for metric in report.sales_by_day
            )
        )

        products_sheet = workbook.create_sheet("Productos m\u00e1s vendidos")
        self._append_table(
            products_sheet,
            ["Producto", "Cantidad vendida", "Total estimado", "Porcentaje sobre ventas"],
            (
                [
                    metric.product,
                    metric.quantity_sold,
                    metric.estimated_sales,
                    self._as_percentage_value(metric.sales_percentage),
                ]
                for metric in report.top_products
            )
        )

        hours_sheet = workbook.create_sheet("Horas pico")
        self._append_table(
            hours_sheet,
            ["Hora", "Total de pedidos", "Total vendido"],
            ([metric.hour, metric.total_orders, metric.total_sales] for metric in report.orders_by_hour)
        )

        status_total = sum(metric.total for metric in report.orders_by_status)
        status_sheet = workbook.create_sheet("\u00d3rdenes por estado")
        self._append_table(
            status_sheet,
            ["Estado", "Total", "Porcentaje"],
            (
                [metric.status, metric.total, (metric.total / status_total) if status_total > 0 else 0]
                for metric in report.orders_by_status
            )
        )

        payments_sheet = workbook.create_sheet("Pagos")
        self._build_payments_sheet(payments_sheet, report)

        catalog_sheet = workbook.create_sheet("Cat\u00e1logo")
        self._append_table(
            catalog_sheet,
            ["M\u00e9trica", "Valor"],
            [
                ["Total de productos", report.catalog.total_products],
                ["Productos disponibles", report.catalog.available_products],
                ["Productos pausados", report.catalog.paused_products],
                ["Productos simples", report.catalog.simple_products],
                ["Productos personalizables", report.catalog.customizable_products],
                ["Productos destacados", report.catalog.featured_products],
                ["Total de categor\u00edas", report.catalog.total_categories],
                ["Categor\u00edas activas", report.catalog.active_categories],
            ]
        )

        self._format_workbook(workbook)
        self._apply_analytics_number_formats(workbook)

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output

    def _build_summary_sheet(
        self,
        sheet,
        report: OwnerAnalyticsReport,
        from_date: date,
        to_date: date
    ) -> None:
        sheet.append(["Reporte Tap2Eat"])
        sheet.append(["Restaurante", report.metadata.restaurant_name])
        sheet.append(["RFC", report.metadata.restaurant_rfc or "No disponible"])
        sheet.append(["Rango de fechas", f"{from_date.isoformat()} a {to_date.isoformat()}"])
        sheet.append(["Sucursal", report.metadata.branch_name])
        sheet.append(["Fecha de generaci\u00f3n", datetime.now().strftime("%Y-%m-%d %H:%M")])
        if report.metadata.branch_filter_note:
            sheet.append(["Nota de sucursal", report.metadata.branch_filter_note])
        sheet.append([])
        sheet.append(["M\u00e9trica", "Valor"])
        sheet.append(["Ventas totales", report.summary.total_sales])
        sheet.append(["Ventas entregadas", report.summary.delivered_sales])
        sheet.append(["\u00d3rdenes totales", report.summary.total_orders])
        sheet.append(["\u00d3rdenes entregadas", report.summary.delivered_orders])
        sheet.append(["\u00d3rdenes canceladas", report.summary.cancelled_orders])
        sheet.append(["Porcentaje de cancelaci\u00f3n", report.summary.cancellation_rate / 100])
        sheet.append(["Ticket promedio", report.summary.average_ticket])
        sheet.append(["Total de productos vendidos", report.summary.total_products_sold])
        sheet.append([
            "M\u00e9todo de pago predominante",
            report.summary.predominant_payment_method or "No disponible"
        ])
        sheet.append([
            "Nota de pagos",
            report.payment_summary.message
            if report.payment_summary.message
            else ""
        ])

    def _build_payments_sheet(self, sheet, report: OwnerAnalyticsReport) -> None:
        if report.payment_summary.message and report.payment_summary.total_payments == 0:
            sheet.append(["Pagos"])
            sheet.append([report.payment_summary.message])
            return

        if not report.payment_summary.available:
            sheet.append(["Pagos"])
            sheet.append([
                report.payment_summary.message
                or "No se pudo cargar la informaci\u00f3n de pagos."
            ])
            return

        self._append_table(
            sheet,
            ["M\u00e9trica", "Valor"],
            [
                ["Total aprobado", report.payment_summary.total_approved],
                ["Efectivo", report.payment_summary.cash],
                ["Online", report.payment_summary.online],
                ["Pagos aprobados", report.payment_summary.approved_payments],
                ["Pagos pendientes", report.payment_summary.pending_payments],
                ["Rechazados o cancelados", report.payment_summary.rejected_or_cancelled],
                ["Cantidad recibida en efectivo", report.payment_summary.cash_amount_received],
                ["Cambio entregado", report.payment_summary.cash_change_amount],
            ]
        )

    def _append_table(self, sheet, headers: list[str], rows: Iterable[list]) -> None:
        sheet.append(headers)
        row_count = 0
        for row in rows:
            sheet.append(row)
            row_count += 1

        if row_count == 0:
            sheet.append(["No hay datos disponibles"])

    def _format_workbook(self, workbook: Workbook) -> None:
        for sheet in workbook.worksheets:
            self._format_sheet(sheet)

    def _format_sheet(self, sheet) -> None:
        if sheet.max_row == 0:
            return

        header_row = self._detect_header_row(sheet)
        if header_row:
            sheet.freeze_panes = f"A{header_row + 1}"
            sheet.auto_filter.ref = (
                f"A{header_row}:{get_column_letter(sheet.max_column)}{sheet.max_row}"
            )

            for cell in sheet[header_row]:
                cell.fill = self._HEADER_FILL
                cell.font = self._HEADER_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center")

        if sheet.title == "Resumen ejecutivo":
            sheet["A1"].font = self._TITLE_FONT
            for cell in sheet["A"]:
                if cell.row in (2, 3, 4, 5, 6, 7):
                    cell.font = self._SUBTITLE_FONT

        for row in sheet.iter_rows():
            for cell in row:
                cell.border = self._BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=True)

        self._auto_fit_columns(sheet)

    def _detect_header_row(self, sheet) -> int | None:
        for row_number in range(1, min(sheet.max_row, 12) + 1):
            values = [cell.value for cell in sheet[row_number]]
            if values and values[0] in ("M\u00e9trica", "Folio pedido", "Fecha", "Producto", "Hora", "Estado"):
                return row_number
            if values and values[0] == "Pagos" and sheet.max_row == 1:
                return row_number

        return 1 if sheet.max_column > 1 else None

    def _auto_fit_columns(self, sheet) -> None:
        for column_cells in sheet.columns:
            column_letter = get_column_letter(column_cells[0].column)
            width = 12
            for cell in column_cells:
                if cell.value is None:
                    continue
                width = max(width, min(len(str(cell.value)) + 2, 42))
            sheet.column_dimensions[column_letter].width = width

    def _apply_analytics_number_formats(self, workbook: Workbook) -> None:
        currency_columns = {
            "Detalle de pedidos": ("I", "J", "K", "L"),
            "Detalle de productos vendidos": ("F", "G"),
            "Ventas por d\u00eda": ("B", "D"),
            "Productos m\u00e1s vendidos": ("C",),
            "Horas pico": ("C",),
            "Pagos": ("B",),
        }
        percentage_columns = {
            "Productos m\u00e1s vendidos": ("D",),
            "\u00d3rdenes por estado": ("C",),
        }

        for sheet_name, columns in currency_columns.items():
            if sheet_name not in workbook.sheetnames:
                continue
            for column in columns:
                self._format_column(workbook[sheet_name], column, self._CURRENCY_FORMAT)

        for sheet_name, columns in percentage_columns.items():
            if sheet_name not in workbook.sheetnames:
                continue
            for column in columns:
                self._format_column(workbook[sheet_name], column, self._PERCENT_FORMAT)

        summary_sheet = workbook["Resumen ejecutivo"]
        for row in range(9, summary_sheet.max_row + 1):
            label = summary_sheet[f"A{row}"].value
            if label in ("Ventas totales", "Ventas entregadas", "Ticket promedio"):
                summary_sheet[f"B{row}"].number_format = self._CURRENCY_FORMAT
            if label == "Porcentaje de cancelaci\u00f3n":
                summary_sheet[f"B{row}"].number_format = self._PERCENT_FORMAT

    def _format_column(self, sheet, column: str, number_format: str) -> None:
        for cell in sheet[column][1:]:
            if isinstance(cell.value, (int, float)):
                cell.number_format = number_format

    def _as_percentage_value(self, percentage: float | None) -> float | None:
        if percentage is None:
            return None

        return percentage / 100
