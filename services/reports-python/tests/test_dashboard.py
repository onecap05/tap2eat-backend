import unittest
from io import BytesIO
from unittest.mock import AsyncMock, MagicMock, patch

import httpx
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.clients.catalog_client import CatalogClient
from app.clients.finance_client import FinanceClient
from app.clients.order_client import OrderClient
from app.core.config import settings
from app.main import app
from app.schemas.dashboard_schema import (
    CatalogReport,
    OrdersReport,
    OwnerAnalyticsReport,
    OwnerAnalyticsMetadata,
    OwnerAnalyticsSummary,
    OwnerDashboardResponse,
    PaymentSummary,
)
from app.services.dashboard_service import DashboardService
from app.services.excel_export_service import ExcelExportService


def build_dashboard_response() -> OwnerDashboardResponse:
    return OwnerDashboardResponse(
        restaurant_id="restaurant-1",
        orders=OrdersReport(
            total_orders=1,
            sales_order_count=1,
            total_sales=25.5,
            delivered_sales=25.5,
            average_ticket=25.5,
            created_orders=0,
            accepted_orders=0,
            preparing_orders=0,
            ready_orders=0,
            delivered_orders=1,
            cancelled_orders=0,
            orders_by_status=[{"status": "DELIVERED", "total": 1}],
        ),
        catalog=CatalogReport(
            total_products=1,
            available_products=1,
            paused_products=0,
            simple_products=1,
            customizable_products=0,
            featured_products=1,
            products_with_custom_schedule=0,
            total_categories=1,
            active_categories=1,
            products_by_type=[{"productType": "SIMPLE", "total": 1}],
            products_by_status=[{"status": "AVAILABLE", "total": 1}],
        ),
    )


def build_analytics_response() -> OwnerAnalyticsReport:
    dashboard = build_dashboard_response()

    return OwnerAnalyticsReport(
        restaurant_id="restaurant-1",
        metadata=OwnerAnalyticsMetadata(
            restaurant_name="Restaurante Demo",
            restaurant_rfc="TAP260520ABC",
            branch_id=None,
            branch_name="Todas las sucursales",
            branch_filter_note=None,
        ),
        summary=OwnerAnalyticsSummary(
            total_sales=25.5,
            total_orders=1,
            average_ticket=25.5,
            delivered_orders=1,
            cancelled_orders=0,
            cancellation_rate=0,
            delivered_sales=25.5,
            total_products_sold=2,
            predominant_payment_method=None,
        ),
        sales_by_day=[
            {
                "date": "2026-01-01",
                "totalSales": 25.5,
                "totalOrders": 1,
                "averageTicket": 25.5,
            }
        ],
        top_products=[
            {
                "product": "Taco",
                "quantitySold": 2,
                "estimatedSales": 25.5,
                "salesPercentage": 100,
            }
        ],
        orders_by_hour=[{"hour": "12:00", "totalOrders": 1, "totalSales": 25.5}],
        orders_by_status=[{"status": "DELIVERED", "total": 1}],
        payment_summary=PaymentSummary(
            available=True,
            total_payments=1,
            total_approved=0,
            cash=0,
            online=0,
            pending=0,
            rejected_or_cancelled=0,
            approved_payments=0,
            pending_payments=0,
            cash_amount_received=0,
            cash_change_amount=0,
            message=None,
        ),
        order_details=[
            {
                "orderFolio": "ORD-1",
                "date": "2026-01-01",
                "time": "12:30",
                "branch": "Sucursal Centro",
                "customer": "Ana",
                "orderStatus": "DELIVERED",
                "paymentMethod": "No disponible",
                "paymentStatus": "No disponible",
                "subtotal": None,
                "total": 25.5,
                "amountReceived": None,
                "changeAmount": None,
                "paymentProvider": "No disponible",
                "paymentReference": "No disponible",
                "itemsCount": 2,
                "notes": "",
            }
        ],
        sold_product_details=[
            {
                "orderFolio": "ORD-1",
                "date": "2026-01-01",
                "branch": "Sucursal Centro",
                "product": "Taco",
                "quantity": 2,
                "unitPrice": 12.75,
                "productTotal": 25.5,
                "modifiers": "",
                "orderStatus": "DELIVERED",
            }
        ],
        catalog=dashboard.catalog,
    )


class DashboardRoutesTest(unittest.TestCase):
    def setUp(self) -> None:
        self.client = TestClient(app)

    def test_health_check_returns_service_status(self) -> None:
        response = self.client.get("/api/reports/health")

        self.assertEqual(200, response.status_code)
        self.assertEqual(
            {"status": "UP", "service": "report-service"},
            response.json(),
        )

    def test_owner_dashboard_returns_response_and_forwards_inputs(self) -> None:
        service = MagicMock()
        service.get_owner_dashboard = AsyncMock(return_value=build_dashboard_response())

        with patch("app.api.dashboard_routes.DashboardService", return_value=service):
            response = self.client.get(
                "/api/reports/dashboard/owner/restaurant-1",
                params={"from": "2026-01-01", "to": "2026-01-31"},
                headers={"Authorization": "Bearer token"},
            )

        self.assertEqual(200, response.status_code)
        body = response.json()
        self.assertEqual("restaurant-1", body["restaurantId"])
        self.assertEqual(1, body["orders"]["salesOrderCount"])
        self.assertEqual(1, body["catalog"]["availableProducts"])
        service.get_owner_dashboard.assert_awaited_once_with(
            restaurant_id="restaurant-1",
            from_date="2026-01-01",
            to_date="2026-01-31",
            authorization_header="Bearer token",
        )

    def test_owner_dashboard_converts_http_status_error_to_bad_gateway(self) -> None:
        request = httpx.Request("GET", "http://orders")
        upstream_response = httpx.Response(503, text="service unavailable", request=request)
        service = MagicMock()
        service.get_owner_dashboard = AsyncMock(
            side_effect=httpx.HTTPStatusError(
                "upstream error",
                request=request,
                response=upstream_response,
            )
        )

        with patch("app.api.dashboard_routes.DashboardService", return_value=service):
            response = self.client.get("/api/reports/dashboard/owner/restaurant-1")

        self.assertEqual(502, response.status_code)
        self.assertEqual(
            "Required service returned 503: service unavailable",
            response.json()["detail"],
        )

    def test_owner_dashboard_converts_http_error_to_bad_gateway(self) -> None:
        service = MagicMock()
        service.get_owner_dashboard = AsyncMock(
            side_effect=httpx.ConnectError("connection refused")
        )

        with patch("app.api.dashboard_routes.DashboardService", return_value=service):
            response = self.client.get("/api/reports/dashboard/owner/restaurant-1")

        self.assertEqual(502, response.status_code)
        self.assertEqual(
            "Could not get data from required services: connection refused",
            response.json()["detail"],
        )

    def test_owner_analytics_returns_response_and_forwards_inputs(self) -> None:
        service = MagicMock()
        service.get_owner_analytics = AsyncMock(return_value=build_analytics_response())

        with patch("app.api.dashboard_routes.DashboardService", return_value=service):
            response = self.client.get(
                "/api/reports/dashboard/owner/restaurant-1/analytics",
                params={"from": "2026-01-01", "to": "2026-01-31", "branchId": "branch-1"},
                headers={"Authorization": "Bearer token"},
            )

        self.assertEqual(200, response.status_code)
        body = response.json()
        self.assertEqual("restaurant-1", body["restaurantId"])
        self.assertEqual(25.5, body["summary"]["totalSales"])
        self.assertEqual("Taco", body["topProducts"][0]["product"])
        service.get_owner_analytics.assert_awaited_once_with(
            restaurant_id="restaurant-1",
            from_date="2026-01-01",
            to_date="2026-01-31",
            authorization_header="Bearer token",
            branch_id="branch-1",
        )

    def test_owner_analytics_rejects_invalid_date_range(self) -> None:
        response = self.client.get(
            "/api/reports/dashboard/owner/restaurant-1/analytics",
            params={"from": "2026-02-01", "to": "2026-01-01"},
        )

        self.assertEqual(400, response.status_code)

    def test_owner_analytics_export_forwards_branch_id(self) -> None:
        service = MagicMock()
        service.get_owner_analytics = AsyncMock(return_value=build_analytics_response())
        excel_service = MagicMock()
        excel_service.build_owner_analytics_workbook.return_value = BytesIO(b"xlsx")

        with patch("app.api.dashboard_routes.DashboardService", return_value=service), \
                patch("app.api.dashboard_routes.ExcelExportService", return_value=excel_service):
            response = self.client.get(
                "/api/reports/dashboard/owner/restaurant-1/analytics/export",
                params={"from": "2026-01-01", "to": "2026-01-31", "branchId": "branch-1"},
                headers={"Authorization": "Bearer token"},
            )

        self.assertEqual(200, response.status_code)
        service.get_owner_analytics.assert_awaited_once_with(
            restaurant_id="restaurant-1",
            from_date="2026-01-01",
            to_date="2026-01-31",
            authorization_header="Bearer token",
            branch_id="branch-1",
        )


class DashboardServiceTest(unittest.IsolatedAsyncioTestCase):
    async def test_get_owner_dashboard_aggregates_orders_and_catalog_data(self) -> None:
        service = DashboardService()
        service.order_client = MagicMock()
        service.catalog_client = MagicMock()

        service.order_client.get_orders_by_restaurant = AsyncMock(
            return_value=[
                {"status": "created", "total": "10.50"},
                {"status": "DELIVERED", "total": "20.25"},
                {"status": "cancelled", "total": "999.99"},
                {"status": "", "total": "invalid"},
                {"total": None},
            ]
        )
        service.catalog_client.get_products_by_restaurant = AsyncMock(
            return_value=[
                {
                    "active": True,
                    "productType": "simple",
                    "featured": True,
                    "availability": {
                        "status": "available",
                        "weeklySchedule": [{"day": "MONDAY"}],
                    },
                },
                {
                    "isActive": True,
                    "productType": "customizable",
                    "availability": {"status": "OUT_OF_STOCK"},
                },
                {
                    "active": False,
                    "productType": "simple",
                    "availability": {"status": "AVAILABLE"},
                },
                {
                    "productType": None,
                    "availability": "always",
                },
            ]
        )
        service.catalog_client.get_categories_by_restaurant = AsyncMock(
            return_value=[
                {"active": True},
                {"isActive": False},
                {},
            ]
        )

        result = await service.get_owner_dashboard(
            restaurant_id="restaurant-1",
            from_date="2026-01-01",
            to_date="2026-01-31",
            authorization_header="Bearer token",
        )

        self.assertEqual("restaurant-1", result.restaurant_id)
        self.assertEqual(5, result.orders.total_orders)
        self.assertEqual(4, result.orders.sales_order_count)
        self.assertEqual(30.75, result.orders.total_sales)
        self.assertEqual(20.25, result.orders.delivered_sales)
        self.assertEqual(7.69, result.orders.average_ticket)
        self.assertEqual(1, result.orders.created_orders)
        self.assertEqual(1, result.orders.delivered_orders)
        self.assertEqual(1, result.orders.cancelled_orders)
        self.assertEqual(
            {"CREATED": 1, "DELIVERED": 1, "CANCELLED": 1, "UNKNOWN": 2},
            {metric.status: metric.total for metric in result.orders.orders_by_status},
        )
        self.assertEqual(3, result.catalog.total_products)
        self.assertEqual(1, result.catalog.available_products)
        self.assertEqual(1, result.catalog.paused_products)
        self.assertEqual(1, result.catalog.simple_products)
        self.assertEqual(1, result.catalog.customizable_products)
        self.assertEqual(1, result.catalog.featured_products)
        self.assertEqual(1, result.catalog.products_with_custom_schedule)
        self.assertEqual(3, result.catalog.total_categories)
        self.assertEqual(2, result.catalog.active_categories)
        self.assertEqual(
            {"SIMPLE": 1, "CUSTOMIZABLE": 1, "UNKNOWN": 1},
            {
                metric.product_type: metric.total
                for metric in result.catalog.products_by_type
            },
        )
        self.assertEqual(
            {"AVAILABLE": 1, "OUT_OF_STOCK": 1, "UNKNOWN": 1},
            {metric.status: metric.total for metric in result.catalog.products_by_status},
        )
        service.order_client.get_orders_by_restaurant.assert_awaited_once_with(
            restaurant_id="restaurant-1",
            from_date="2026-01-01",
            to_date="2026-01-31",
        )
        service.catalog_client.get_products_by_restaurant.assert_awaited_once_with(
            restaurant_id="restaurant-1",
            authorization_header="Bearer token",
        )
        service.catalog_client.get_categories_by_restaurant.assert_awaited_once_with(
            restaurant_id="restaurant-1",
            authorization_header="Bearer token",
        )

    def test_empty_orders_report_uses_zero_totals(self) -> None:
        service = DashboardService()

        result = service._build_orders_report([])

        self.assertEqual(0, result.total_orders)
        self.assertEqual(0, result.sales_order_count)
        self.assertEqual(0.0, result.total_sales)
        self.assertEqual(0.0, result.delivered_sales)
        self.assertEqual(0.0, result.average_ticket)
        self.assertEqual([], result.orders_by_status)

    async def test_get_owner_analytics_calculates_sales_products_and_peak_hours(self) -> None:
        service = DashboardService()
        service.order_client = MagicMock()
        service.catalog_client = MagicMock()
        service.finance_client = MagicMock()
        service.order_client.get_orders_by_restaurant = AsyncMock(
            return_value=[
                {
                    "id": "order-1",
                    "status": "DELIVERED",
                    "total": "120.00",
                    "branchId": "branch-1",
                    "branchName": "Sucursal Centro",
                    "createdAt": "2026-01-01T12:30:00Z",
                    "items": [
                        {
                            "productNameSnapshot": "Taco",
                            "quantity": 2,
                            "subtotal": "80.00",
                        },
                        {
                            "productNameSnapshot": "Agua",
                            "quantity": 1,
                            "subtotal": "40.00",
                        },
                    ],
                },
                {
                    "id": "order-2",
                    "status": "CANCELLED",
                    "total": "50.00",
                    "branchId": "branch-2",
                    "createdAt": "2026-01-01T13:00:00Z",
                    "items": [
                        {
                            "productNameSnapshot": "Taco",
                            "quantity": 10,
                            "subtotal": "500.00",
                        }
                    ],
                },
                {
                    "id": "order-3",
                    "status": "READY",
                    "total": "30.00",
                    "branchId": "branch-1",
                    "branchName": "Sucursal Centro",
                    "createdAt": "2026-01-02T12:10:00Z",
                    "items": [
                        {
                            "productNameSnapshot": "Taco",
                            "quantity": 1,
                            "subtotal": "30.00",
                        }
                    ],
                },
            ]
        )
        service.catalog_client.get_products_by_restaurant = AsyncMock(return_value=[])
        service.catalog_client.get_categories_by_restaurant = AsyncMock(return_value=[])
        service.catalog_client.get_restaurant_by_id = AsyncMock(
            return_value={"id": "restaurant-1", "name": "Restaurante Demo", "rfc": "TAP260520ABC"}
        )
        service.catalog_client.get_branches_by_restaurant = AsyncMock(
            return_value=[{"id": "branch-1", "name": "Sucursal Centro"}]
        )
        service.finance_client.get_payments_by_restaurant = AsyncMock(
            return_value=[
                {
                    "orderId": "order-1",
                    "branchId": "branch-1",
                    "amount": "120.00",
                    "status": "Approved",
                    "provider": "CASH",
                    "amountReceived": "150.00",
                    "changeAmount": "30.00",
                    "approvedAt": "2026-01-01T13:00:00Z",
                },
                {
                    "orderId": "order-3",
                    "branchId": "branch-1",
                    "amount": "30.00",
                    "status": "Pending",
                    "provider": "PAYPAL",
                    "createdAt": "2026-01-02T12:10:00Z",
                },
            ]
        )

        result = await service.get_owner_analytics("restaurant-1")

        self.assertEqual(3, result.summary.total_orders)
        self.assertEqual(150.0, result.summary.total_sales)
        self.assertEqual(1, result.summary.cancelled_orders)
        self.assertEqual(33.33, result.summary.cancellation_rate)
        self.assertEqual(4, result.summary.total_products_sold)
        self.assertEqual(
            [("2026-01-01", 120.0, 1, 120.0), ("2026-01-02", 30.0, 1, 30.0)],
            [
                (
                    metric.date,
                    metric.total_sales,
                    metric.total_orders,
                    metric.average_ticket,
                )
                for metric in result.sales_by_day
            ],
        )
        self.assertEqual("Taco", result.top_products[0].product)
        self.assertEqual(3, result.top_products[0].quantity_sold)
        self.assertEqual(2, result.orders_by_hour[0].total_orders)
        self.assertEqual(150.0, result.orders_by_hour[0].total_sales)
        self.assertEqual(3, len(result.order_details))
        self.assertEqual(3, len(result.sold_product_details))
        self.assertTrue(result.payment_summary.available)
        self.assertEqual(2, result.payment_summary.total_payments)
        self.assertEqual(120.0, result.payment_summary.cash)
        self.assertEqual(150.0, result.payment_summary.cash_amount_received)
        self.assertEqual("Efectivo", result.summary.predominant_payment_method)
        self.assertEqual("TAP260520ABC", result.metadata.restaurant_rfc)

    async def test_get_owner_analytics_filters_orders_and_payments_by_branch(self) -> None:
        service = DashboardService()
        service.order_client = MagicMock()
        service.catalog_client = MagicMock()
        service.finance_client = MagicMock()
        service.order_client.get_orders_by_restaurant = AsyncMock(
            return_value=[
                {
                    "id": "order-1",
                    "branchId": "branch-1",
                    "branchName": "Sucursal Centro",
                    "status": "DELIVERED",
                    "total": "100.00",
                    "createdAt": "2026-01-01T10:00:00Z",
                    "items": [{"productNameSnapshot": "Taco", "quantity": 1, "subtotal": "100.00"}],
                },
                {
                    "id": "order-2",
                    "branchId": "branch-2",
                    "branchName": "Sucursal Norte",
                    "status": "DELIVERED",
                    "total": "200.00",
                    "createdAt": "2026-01-01T11:00:00Z",
                    "items": [{"productNameSnapshot": "Agua", "quantity": 1, "subtotal": "200.00"}],
                },
            ]
        )
        service.catalog_client.get_products_by_restaurant = AsyncMock(return_value=[])
        service.catalog_client.get_categories_by_restaurant = AsyncMock(return_value=[])
        service.catalog_client.get_restaurant_by_id = AsyncMock(return_value={"name": "Demo"})
        service.catalog_client.get_branches_by_restaurant = AsyncMock(
            return_value=[{"id": "branch-1", "name": "Sucursal Centro"}]
        )
        service.finance_client.get_payments_by_restaurant = AsyncMock(
            return_value=[
                {
                    "orderId": "order-1",
                    "branchId": "branch-1",
                    "amount": "100.00",
                    "status": "Approved",
                    "provider": "PAYPAL",
                    "approvedAt": "2026-01-01T10:30:00Z",
                },
                {
                    "orderId": "order-2",
                    "branchId": "branch-2",
                    "amount": "200.00",
                    "status": "Approved",
                    "provider": "CASH",
                    "approvedAt": "2026-01-01T11:30:00Z",
                },
            ]
        )

        result = await service.get_owner_analytics(
            "restaurant-1",
            from_date="2026-01-01T00:00:00",
            to_date="2026-01-01T23:59:59",
            branch_id="branch-1",
        )

        self.assertEqual(1, result.summary.total_orders)
        self.assertEqual(100.0, result.summary.total_sales)
        self.assertEqual("Sucursal Centro", result.metadata.branch_name)
        self.assertEqual(1, result.payment_summary.total_payments)
        self.assertEqual(100.0, result.payment_summary.online)

    async def test_get_owner_analytics_handles_finance_failure_without_breaking_report(self) -> None:
        service = DashboardService()
        service.order_client = MagicMock()
        service.catalog_client = MagicMock()
        service.finance_client = MagicMock()
        service.order_client.get_orders_by_restaurant = AsyncMock(return_value=[])
        service.catalog_client.get_products_by_restaurant = AsyncMock(return_value=[])
        service.catalog_client.get_categories_by_restaurant = AsyncMock(return_value=[])
        service.catalog_client.get_restaurant_by_id = AsyncMock(return_value={"name": "Demo"})
        service.catalog_client.get_branches_by_restaurant = AsyncMock(return_value=[])
        service.finance_client.get_payments_by_restaurant = AsyncMock(
            side_effect=httpx.ConnectError("finance down")
        )

        result = await service.get_owner_analytics("restaurant-1")

        self.assertEqual(0, result.summary.total_orders)
        self.assertFalse(result.payment_summary.available)
        self.assertEqual(
            "No se pudo cargar la información de pagos.",
            result.payment_summary.message,
        )


class ExcelExportServiceTest(unittest.TestCase):
    def test_build_owner_analytics_workbook_creates_expected_sheets(self) -> None:
        service = ExcelExportService()

        workbook_file = service.build_owner_analytics_workbook(
            report=build_analytics_response(),
            from_date=__import__("datetime").date(2026, 1, 1),
            to_date=__import__("datetime").date(2026, 1, 31),
        )

        workbook = load_workbook(BytesIO(workbook_file.getvalue()))

        self.assertEqual(
            [
                "Resumen ejecutivo",
                "Detalle de pedidos",
                "Detalle de productos vendidos",
                "Ventas por día",
                "Productos más vendidos",
                "Horas pico",
                "Órdenes por estado",
                "Pagos",
                "Catálogo",
            ],
            workbook.sheetnames,
        )
        self.assertEqual("Restaurante Demo", workbook["Resumen ejecutivo"]["B2"].value)
        self.assertEqual("TAP260520ABC", workbook["Resumen ejecutivo"]["B3"].value)
        self.assertEqual("Todas las sucursales", workbook["Resumen ejecutivo"]["B5"].value)

    def test_build_owner_analytics_workbook_includes_order_details(self) -> None:
        service = ExcelExportService()

        workbook_file = service.build_owner_analytics_workbook(
            report=build_analytics_response(),
            from_date=__import__("datetime").date(2026, 1, 1),
            to_date=__import__("datetime").date(2026, 1, 31),
        )

        workbook = load_workbook(BytesIO(workbook_file.getvalue()))
        sheet = workbook["Detalle de pedidos"]

        self.assertEqual("Folio pedido", sheet["A1"].value)
        self.assertEqual("Total", sheet["J1"].value)
        self.assertEqual("ORD-1", sheet["A2"].value)
        self.assertEqual(25.5, sheet["J2"].value)
        self.assertEqual('"$"#,##0.00', sheet["J2"].number_format)
        self.assertEqual("A2", sheet.freeze_panes)

    def test_build_owner_analytics_workbook_includes_sold_product_details(self) -> None:
        service = ExcelExportService()

        workbook_file = service.build_owner_analytics_workbook(
            report=build_analytics_response(),
            from_date=__import__("datetime").date(2026, 1, 1),
            to_date=__import__("datetime").date(2026, 1, 31),
        )

        workbook = load_workbook(BytesIO(workbook_file.getvalue()))
        sheet = workbook["Detalle de productos vendidos"]

        self.assertEqual("Producto", sheet["D1"].value)
        self.assertEqual("Taco", sheet["D2"].value)
        self.assertEqual(2, sheet["E2"].value)
        self.assertEqual(12.75, sheet["F2"].value)
        self.assertEqual('"$"#,##0.00', sheet["F2"].number_format)

    def test_build_owner_analytics_workbook_handles_empty_orders(self) -> None:
        service = ExcelExportService()
        report = build_analytics_response()
        report.summary.total_orders = 0
        report.summary.total_sales = 0
        report.summary.delivered_sales = 0
        report.summary.average_ticket = 0
        report.summary.delivered_orders = 0
        report.summary.cancelled_orders = 0
        report.summary.cancellation_rate = 0
        report.summary.total_products_sold = 0
        report.sales_by_day = []
        report.top_products = []
        report.orders_by_hour = []
        report.orders_by_status = []
        report.order_details = []
        report.sold_product_details = []
        report.payment_summary.available = True
        report.payment_summary.total_payments = 0
        report.payment_summary.total_approved = 0
        report.payment_summary.cash = 0
        report.payment_summary.online = 0
        report.payment_summary.pending = 0
        report.payment_summary.rejected_or_cancelled = 0
        report.payment_summary.approved_payments = 0
        report.payment_summary.pending_payments = 0
        report.payment_summary.cash_amount_received = 0
        report.payment_summary.cash_change_amount = 0
        report.payment_summary.message = "No hay pagos registrados para este rango."

        workbook_file = service.build_owner_analytics_workbook(
            report=report,
            from_date=__import__("datetime").date(2026, 1, 1),
            to_date=__import__("datetime").date(2026, 1, 31),
        )

        workbook = load_workbook(BytesIO(workbook_file.getvalue()))

        self.assertEqual(
            "No hay datos disponibles",
            workbook["Detalle de pedidos"]["A2"].value,
        )
        self.assertEqual(
            "No hay datos disponibles",
            workbook["Detalle de productos vendidos"]["A2"].value,
        )
        self.assertEqual(
            "No hay pagos registrados para este rango.",
            workbook["Pagos"]["A2"].value,
        )


class HttpClientsTest(unittest.IsolatedAsyncioTestCase):
    async def test_order_client_sends_internal_token_and_date_filters(self) -> None:
        response = MagicMock()
        response.json.return_value = [{"id": "order-1"}]
        http_client = MagicMock()
        http_client.get = AsyncMock(return_value=response)
        context_manager = MagicMock()
        context_manager.__aenter__ = AsyncMock(return_value=http_client)
        context_manager.__aexit__ = AsyncMock(return_value=None)

        with patch(
            "app.clients.order_client.httpx.AsyncClient",
            return_value=context_manager,
        ) as async_client:
            client = OrderClient()
            client.base_url = "http://orders"
            client.timeout = 3.5

            result = await client.get_orders_by_restaurant(
                restaurant_id="restaurant-1",
                from_date="2026-01-01",
                to_date="2026-01-31",
            )

        self.assertEqual([{"id": "order-1"}], result)
        async_client.assert_called_once_with(timeout=3.5)
        http_client.get.assert_awaited_once_with(
            "http://orders/api/internal/orders/restaurant/restaurant-1",
            params={"from": "2026-01-01", "to": "2026-01-31"},
            headers={"X-Internal-Service-Token": settings.internal_service_token},
        )
        response.raise_for_status.assert_called_once_with()

    async def test_order_client_returns_empty_list_for_non_list_body(self) -> None:
        response = MagicMock()
        response.json.return_value = {"items": []}
        http_client = MagicMock()
        http_client.get = AsyncMock(return_value=response)
        context_manager = MagicMock()
        context_manager.__aenter__ = AsyncMock(return_value=http_client)
        context_manager.__aexit__ = AsyncMock(return_value=None)

        with patch("app.clients.order_client.httpx.AsyncClient", return_value=context_manager):
            client = OrderClient()
            client.base_url = "http://orders"

            result = await client.get_orders_by_restaurant("restaurant-1")

        self.assertEqual([], result)
        http_client.get.assert_awaited_once_with(
            "http://orders/api/internal/orders/restaurant/restaurant-1",
            params={},
            headers={"X-Internal-Service-Token": settings.internal_service_token},
        )

    async def test_catalog_client_sends_authorization_header(self) -> None:
        response = MagicMock()
        response.json.return_value = [{"id": "product-1"}]
        http_client = MagicMock()
        http_client.get = AsyncMock(return_value=response)
        context_manager = MagicMock()
        context_manager.__aenter__ = AsyncMock(return_value=http_client)
        context_manager.__aexit__ = AsyncMock(return_value=None)

        with patch(
            "app.clients.catalog_client.httpx.AsyncClient",
            return_value=context_manager,
        ):
            client = CatalogClient()
            client.base_url = "http://catalog"

            result = await client.get_products_by_restaurant(
                restaurant_id="restaurant-1",
                authorization_header="Bearer token",
            )

        self.assertEqual([{"id": "product-1"}], result)
        http_client.get.assert_awaited_once_with(
            "http://catalog/api/products/restaurant/restaurant-1",
            headers={"Authorization": "Bearer token"},
        )
        response.raise_for_status.assert_called_once_with()

    async def test_catalog_client_returns_empty_list_for_non_list_products_body(self) -> None:
        response = MagicMock()
        response.json.return_value = {"items": []}
        http_client = MagicMock()
        http_client.get = AsyncMock(return_value=response)
        context_manager = MagicMock()
        context_manager.__aenter__ = AsyncMock(return_value=http_client)
        context_manager.__aexit__ = AsyncMock(return_value=None)

        with patch("app.clients.catalog_client.httpx.AsyncClient", return_value=context_manager):
            client = CatalogClient()
            client.base_url = "http://catalog"

            result = await client.get_products_by_restaurant("restaurant-1")

        self.assertEqual([], result)
        http_client.get.assert_awaited_once_with(
            "http://catalog/api/products/restaurant/restaurant-1",
            headers={},
        )

    async def test_catalog_client_returns_categories_list(self) -> None:
        response = MagicMock()
        response.json.return_value = [{"id": "category-1"}]
        http_client = MagicMock()
        http_client.get = AsyncMock(return_value=response)
        context_manager = MagicMock()
        context_manager.__aenter__ = AsyncMock(return_value=http_client)
        context_manager.__aexit__ = AsyncMock(return_value=None)

        with patch("app.clients.catalog_client.httpx.AsyncClient", return_value=context_manager):
            client = CatalogClient()
            client.base_url = "http://catalog"

            result = await client.get_categories_by_restaurant(
                restaurant_id="restaurant-1",
                authorization_header="Bearer token",
            )

        self.assertEqual([{"id": "category-1"}], result)
        http_client.get.assert_awaited_once_with(
            "http://catalog/api/categories/restaurant/restaurant-1",
            headers={"Authorization": "Bearer token"},
        )
        response.raise_for_status.assert_called_once_with()

    async def test_catalog_client_returns_empty_list_for_non_list_categories_body(self) -> None:
        response = MagicMock()
        response.json.return_value = {"items": []}
        http_client = MagicMock()
        http_client.get = AsyncMock(return_value=response)
        context_manager = MagicMock()
        context_manager.__aenter__ = AsyncMock(return_value=http_client)
        context_manager.__aexit__ = AsyncMock(return_value=None)

        with patch("app.clients.catalog_client.httpx.AsyncClient", return_value=context_manager):
            client = CatalogClient()
            client.base_url = "http://catalog"

            result = await client.get_categories_by_restaurant("restaurant-1")

        self.assertEqual([], result)
        http_client.get.assert_awaited_once_with(
            "http://catalog/api/categories/restaurant/restaurant-1",
            headers={},
        )


if __name__ == "__main__":
    unittest.main()
